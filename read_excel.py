import openpyxl
import os

# Caminho do arquivo
file_path = r'ignorar\Inventário_SONEPAR_Consolidado 2026 (6).xlsm'

# Verificar se arquivo existe
if os.path.exists(file_path):
    print(f'✓ Arquivo encontrado')
    
    # Carregar workbook
    wb = openpyxl.load_workbook(file_path)
    
    # 1. Listar todas as abas
    print('=' * 60)
    print('1. TODAS AS ABAS DO ARQUIVO:')
    print('=' * 60)
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f'{i}. {sheet_name}')
    print()
    
    # 2. Verificar se existe a aba 'Inventário SONEPAR'
    target_sheet = 'Inventário SONEPAR'
    if target_sheet in wb.sheetnames:
        print('=' * 60)
        print(f'2. COLUNAS DA ABA "{target_sheet}":')
        print('=' * 60)
        ws = wb[target_sheet]
        
        # Obter headers da primeira linha
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f'Total de colunas preenchidas: {len(headers)}')
        print()
        for i, header in enumerate(headers, 1):
            print(f'{i}. {header}')
    else:
        print(f'⚠ Aba "{target_sheet}" não encontrada no arquivo')
        print(f'Abas disponíveis: {wb.sheetnames}')
else:
    print(f'✗ Arquivo não encontrado em: {file_path}')
